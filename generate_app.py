#!/usr/bin/env python3
"""
Southern 100 Web App Generator
Usage: python generate_app.py
Reads Southern_100_Catalogue_NEW.xlsx and outputs southern100.html
"""

import openpyxl
import json
import os

XLSX_PATH = "Southern_100_Catalogue_NEW.xlsx"
OUTPUT_PATH = "southern100.html"

def extract_data(xlsx_path):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb['Southern 100 Catalogue']
    ws_sp = wb['Seasonal Planner']

    months = ['Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec']

    planner = {}
    for row in ws_sp.iter_rows(min_row=3, values_only=True):
        if row[0] and isinstance(row[0], int):
            dots = {}
            for i, m in enumerate(months):
                val = row[i+2]
                dots[m] = val if val in ['●','↑'] else ''
            planner[row[0]] = dots

    entries = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        if row[0] and isinstance(row[0], int):
            num = row[0]
            entries.append({
                'num': num,
                'act': str(row[1] or ''),
                'cat_id': str(row[2] or ''),
                'name': str(row[3] or ''),
                'type': str(row[4] or ''),
                'ra': str(row[5] or ''),
                'dec': str(row[6] or ''),
                'rotation': str(row[7] or ''),
                'size': str(row[8] or ''),
                'constellation': str(row[9] or ''),
                'transit_alt': str(row[10] or ''),
                'circumpolar': str(row[11] or ''),
                'best_rig': str(row[12] or ''),
                'filter_rec': str(row[13] or ''),
                'best_months': str(row[14] or ''),
                'status': str(row[15] or ''),
                'hours': str(row[16] or ''),
                'notes': str(row[17] or ''),
                'capture_seq': str(row[18] or ''),
                'moon_sep': str(row[19] or ''),
                'planner': planner.get(num, {m: '' for m in months})
            })
    return entries

def build_html(entries):
    data_json = json.dumps(entries, ensure_ascii=False)

    html = f'''<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>The Southern 100</title>
<link rel="preconnect" href="https://fonts.googleapis.com">
<link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
<link href="https://fonts.googleapis.com/css2?family=Libre+Baskerville:ital,wght@0,400;0,700;1,400&family=JetBrains+Mono:wght@300;400;500&display=swap" rel="stylesheet">
<style>
:root {{
  --bg:        #0a0c10;
  --bg2:       #0f1218;
  --bg3:       #151a22;
  --border:    #1e2530;
  --border2:   #2a3545;
  --text:      #c8d4e0;
  --text2:     #6a7f96;
  --text3:     #3d5066;
  --accent:    #4a9eff;
  --accent2:   #1a4a80;
  --gold:      #c8a84b;
  --gold2:     #8a6a20;
  --green:     #3dba7a;
  --green2:    #1a5a38;
  --amber:     #e8a030;
  --amber2:    #7a4a10;
  --red:       #e05050;
  --serif:     'Libre Baskerville', Georgia, serif;
  --mono:      'JetBrains Mono', 'Courier New', monospace;
  --radius:    6px;
  --transition: 180ms ease;
}}

*, *::before, *::after {{ box-sizing: border-box; margin: 0; padding: 0; }}

html {{ font-size: 15px; }}

body {{
  background: var(--bg);
  color: var(--text);
  font-family: var(--serif);
  min-height: 100vh;
  line-height: 1.6;
}}

/* ── HEADER ── */
header {{
  border-bottom: 1px solid var(--border);
  background: var(--bg2);
  padding: 0 1.5rem;
  display: flex;
  align-items: center;
  justify-content: space-between;
  gap: 1rem;
  height: 56px;
  position: sticky;
  top: 0;
  z-index: 100;
}}

.header-title {{
  font-family: var(--mono);
  font-size: 0.72rem;
  font-weight: 500;
  letter-spacing: 0.18em;
  text-transform: uppercase;
  color: var(--gold);
  white-space: nowrap;
}}

.header-sub {{
  font-family: var(--mono);
  font-size: 0.62rem;
  color: var(--text3);
  letter-spacing: 0.1em;
  white-space: nowrap;
}}

.header-left {{ display: flex; flex-direction: column; gap: 2px; }}

/* ── NAV TABS ── */
nav {{
  display: flex;
  gap: 0;
  border-bottom: 1px solid var(--border);
  background: var(--bg2);
  padding: 0 1.5rem;
  position: sticky;
  top: 56px;
  z-index: 99;
}}

.tab {{
  font-family: var(--mono);
  font-size: 0.7rem;
  letter-spacing: 0.12em;
  text-transform: uppercase;
  color: var(--text3);
  padding: 0.7rem 1.2rem;
  cursor: pointer;
  border-bottom: 2px solid transparent;
  transition: color var(--transition), border-color var(--transition);
  user-select: none;
  background: none;
  border-top: none;
  border-left: none;
  border-right: none;
}}

.tab:hover {{ color: var(--text); }}
.tab.active {{ color: var(--gold); border-bottom-color: var(--gold); }}

/* ── VIEWS ── */
.view {{ display: none; }}
.view.active {{ display: block; }}

/* ── CATALOGUE VIEW ── */
.catalogue-controls {{
  padding: 1rem 1.5rem;
  display: flex;
  flex-wrap: wrap;
  gap: 0.6rem;
  align-items: center;
  border-bottom: 1px solid var(--border);
  background: var(--bg2);
  position: sticky;
  top: 97px;
  z-index: 98;
}}

.search-wrap {{
  position: relative;
  flex: 1 1 220px;
  min-width: 180px;
}}

.search-wrap svg {{
  position: absolute;
  left: 0.7rem;
  top: 50%;
  transform: translateY(-50%);
  color: var(--text3);
  pointer-events: none;
}}

input[type=search] {{
  width: 100%;
  background: var(--bg3);
  border: 1px solid var(--border2);
  border-radius: var(--radius);
  color: var(--text);
  font-family: var(--mono);
  font-size: 0.75rem;
  padding: 0.45rem 0.8rem 0.45rem 2.2rem;
  outline: none;
  transition: border-color var(--transition);
}}

input[type=search]:focus {{ border-color: var(--accent2); }}
input[type=search]::placeholder {{ color: var(--text3); }}

select {{
  background: var(--bg3);
  border: 1px solid var(--border2);
  border-radius: var(--radius);
  color: var(--text);
  font-family: var(--mono);
  font-size: 0.72rem;
  padding: 0.45rem 0.7rem;
  outline: none;
  cursor: pointer;
  transition: border-color var(--transition);
}}

select:focus {{ border-color: var(--accent2); }}

.filter-count {{
  font-family: var(--mono);
  font-size: 0.68rem;
  color: var(--text3);
  margin-left: auto;
  white-space: nowrap;
}}

/* ── TABLE ── */
.table-wrap {{
  overflow-x: auto;
  -webkit-overflow-scrolling: touch;
}}

table {{
  width: 100%;
  border-collapse: collapse;
  font-size: 0.82rem;
}}

thead th {{
  font-family: var(--mono);
  font-size: 0.62rem;
  font-weight: 500;
  letter-spacing: 0.12em;
  text-transform: uppercase;
  color: var(--text3);
  padding: 0.6rem 1rem;
  text-align: left;
  border-bottom: 1px solid var(--border);
  background: var(--bg2);
  white-space: nowrap;
  cursor: pointer;
  user-select: none;
}}

thead th:hover {{ color: var(--text2); }}
thead th.sorted {{ color: var(--gold); }}
thead th .sort-arrow {{ margin-left: 4px; opacity: 0.5; }}
thead th.sorted .sort-arrow {{ opacity: 1; }}

tbody tr {{
  border-bottom: 1px solid var(--border);
  transition: background var(--transition);
  cursor: pointer;
}}

tbody tr:hover {{ background: var(--bg3); }}
tbody tr.expanded {{ background: var(--bg3); }}

tbody td {{
  padding: 0.6rem 1rem;
  vertical-align: top;
}}

.col-num {{
  font-family: var(--mono);
  font-size: 0.7rem;
  color: var(--text3);
  width: 42px;
  min-width: 42px;
}}

.col-name {{ min-width: 180px; }}
.col-name strong {{ display: block; color: var(--text); font-weight: 700; font-size: 0.85rem; }}
.col-name span {{ display: block; font-family: var(--mono); font-size: 0.65rem; color: var(--text3); margin-top: 1px; }}

.col-type {{
  font-family: var(--mono);
  font-size: 0.65rem;
  color: var(--text2);
  min-width: 140px;
}}

.col-const {{
  font-family: var(--mono);
  font-size: 0.7rem;
  color: var(--text2);
  white-space: nowrap;
}}

.col-rig {{
  font-family: var(--mono);
  font-size: 0.65rem;
  color: var(--text2);
  min-width: 160px;
}}

.col-months {{
  font-family: var(--mono);
  font-size: 0.68rem;
  color: var(--text2);
  white-space: nowrap;
}}

.col-hours {{
  font-family: var(--mono);
  font-size: 0.68rem;
  color: var(--text2);
  white-space: nowrap;
  text-align: right;
}}

.status-badge {{
  display: inline-block;
  font-family: var(--mono);
  font-size: 0.6rem;
  letter-spacing: 0.08em;
  text-transform: uppercase;
  padding: 2px 7px;
  border-radius: 3px;
  white-space: nowrap;
}}

.status-Complete   {{ background: var(--green2);  color: var(--green);  }}
.status-Improve    {{ background: var(--amber2);   color: var(--amber);  }}
.status-Partial    {{ background: var(--accent2);  color: var(--accent); }}
.status-To-Image   {{ background: var(--bg3);      color: var(--text3);  border: 1px solid var(--border2); }}

/* ── EXPAND ROW ── */
.expand-row td {{
  padding: 0;
  background: var(--bg3);
}}

.expand-content {{
  padding: 1.2rem 1.5rem 1.4rem;
  border-top: 1px solid var(--border2);
  display: grid;
  grid-template-columns: 1fr auto;
  gap: 1.2rem 2rem;
  align-items: start;
}}

.expand-meta {{
  display: flex;
  flex-wrap: wrap;
  gap: 0.4rem 1.5rem;
  margin-bottom: 0.8rem;
}}

.meta-item {{
  font-family: var(--mono);
  font-size: 0.65rem;
}}

.meta-label {{ color: var(--text3); margin-right: 0.3em; }}
.meta-val   {{ color: var(--text2); }}

.notes-text {{
  font-size: 0.88rem;
  color: var(--text);
  line-height: 1.8;
  max-width: 72ch;
}}

.notes-text p {{ margin-bottom: 0.7rem; }}

.expand-actions {{
  display: flex;
  flex-direction: column;
  gap: 0.5rem;
  align-items: flex-end;
}}

.btn {{
  font-family: var(--mono);
  font-size: 0.65rem;
  letter-spacing: 0.1em;
  text-transform: uppercase;
  padding: 0.45rem 1rem;
  border-radius: var(--radius);
  cursor: pointer;
  border: 1px solid transparent;
  transition: all var(--transition);
  white-space: nowrap;
  background: none;
  color: var(--text2);
}}

.btn-primary {{
  background: var(--gold2);
  border-color: var(--gold);
  color: var(--gold);
}}

.btn-primary:hover {{
  background: var(--gold);
  color: var(--bg);
}}

.btn-ghost {{
  border-color: var(--border2);
  color: var(--text3);
}}

.btn-ghost:hover {{
  border-color: var(--text3);
  color: var(--text2);
}}

/* ── PLANNER VIEW ── */
.planner-wrap {{
  padding: 1.5rem;
  overflow-x: auto;
}}

.planner-title {{
  font-family: var(--mono);
  font-size: 0.65rem;
  letter-spacing: 0.15em;
  text-transform: uppercase;
  color: var(--text3);
  margin-bottom: 1.2rem;
}}

.planner-table {{
  border-collapse: collapse;
  font-size: 0.75rem;
  min-width: 700px;
  width: 100%;
}}

.planner-table th {{
  font-family: var(--mono);
  font-size: 0.6rem;
  letter-spacing: 0.1em;
  text-transform: uppercase;
  color: var(--text3);
  padding: 0.4rem 0.5rem;
  text-align: center;
  border-bottom: 1px solid var(--border);
}}

.planner-table th.name-col {{
  text-align: left;
  padding-left: 0;
  min-width: 200px;
}}

.planner-table td {{
  padding: 0.3rem 0.5rem;
  text-align: center;
  border-bottom: 1px solid var(--border);
}}

.planner-table td.name-col {{
  text-align: left;
  padding-left: 0;
}}

.planner-table tr:hover td {{ background: var(--bg3); cursor: pointer; }}

.p-num {{
  font-family: var(--mono);
  font-size: 0.6rem;
  color: var(--text3);
  margin-right: 0.5rem;
}}

.p-name {{
  font-size: 0.75rem;
  color: var(--text);
}}

.p-dot {{
  color: var(--gold);
  font-size: 0.6rem;
  line-height: 1;
}}

.p-rise {{
  color: var(--text3);
  font-size: 0.6rem;
}}

/* ── DETAIL VIEW ── */
.detail-wrap {{
  max-width: 900px;
  margin: 0 auto;
  padding: 2rem 1.5rem 3rem;
}}

.detail-back {{
  font-family: var(--mono);
  font-size: 0.65rem;
  letter-spacing: 0.1em;
  text-transform: uppercase;
  color: var(--text3);
  cursor: pointer;
  display: inline-flex;
  align-items: center;
  gap: 0.4rem;
  margin-bottom: 1.8rem;
  border: none;
  background: none;
  padding: 0;
  transition: color var(--transition);
}}

.detail-back:hover {{ color: var(--text); }}

.detail-header {{
  border-bottom: 1px solid var(--border);
  padding-bottom: 1.2rem;
  margin-bottom: 1.5rem;
}}

.detail-num {{
  font-family: var(--mono);
  font-size: 0.65rem;
  color: var(--gold);
  letter-spacing: 0.15em;
  margin-bottom: 0.3rem;
}}

.detail-name {{
  font-size: 1.6rem;
  font-weight: 700;
  color: var(--text);
  line-height: 1.2;
  margin-bottom: 0.3rem;
}}

.detail-catid {{
  font-family: var(--mono);
  font-size: 0.7rem;
  color: var(--text3);
}}

.detail-meta-grid {{
  display: grid;
  grid-template-columns: repeat(auto-fill, minmax(200px, 1fr));
  gap: 0.8rem 1.5rem;
  margin-bottom: 2rem;
  padding: 1.2rem;
  background: var(--bg2);
  border: 1px solid var(--border);
  border-radius: var(--radius);
}}

.detail-meta-item {{ display: flex; flex-direction: column; gap: 2px; }}
.detail-meta-label {{
  font-family: var(--mono);
  font-size: 0.58rem;
  letter-spacing: 0.12em;
  text-transform: uppercase;
  color: var(--text3);
}}
.detail-meta-val {{
  font-family: var(--mono);
  font-size: 0.75rem;
  color: var(--text);
}}

.detail-section-title {{
  font-family: var(--mono);
  font-size: 0.62rem;
  letter-spacing: 0.15em;
  text-transform: uppercase;
  color: var(--gold);
  margin-bottom: 0.8rem;
  padding-bottom: 0.4rem;
  border-bottom: 1px solid var(--border);
}}

.detail-notes {{
  font-size: 0.92rem;
  color: var(--text);
  line-height: 1.85;
  margin-bottom: 2.5rem;
}}

.detail-notes p {{ margin-bottom: 0.8rem; }}

.detail-capseq {{
  font-family: var(--mono);
  font-size: 0.82rem;
  color: var(--text);
  line-height: 1.9;
  background: var(--bg2);
  border: 1px solid var(--border);
  border-radius: var(--radius);
  padding: 1.4rem 1.6rem;
  margin-bottom: 2rem;
}}

.capseq-segment {{
  margin-bottom: 1.2rem;
  padding-bottom: 1.2rem;
  border-bottom: 1px solid var(--border);
}}

.capseq-segment:last-child {{
  margin-bottom: 0;
  padding-bottom: 0;
  border-bottom: none;
}}

.capseq-label {{
  font-family: var(--mono);
  font-size: 0.65rem;
  letter-spacing: 0.14em;
  text-transform: uppercase;
  color: var(--gold);
  margin-bottom: 0.5rem;
  display: block;
}}

.capseq-text {{
  font-family: var(--mono);
  font-size: 0.82rem;
  color: var(--text);
  line-height: 1.9;
  white-space: pre-wrap;
  word-break: break-word;
}}

/* ── PLANNER MINI IN DETAIL ── */
.detail-planner {{
  display: flex;
  gap: 0.3rem;
  flex-wrap: wrap;
  margin-bottom: 2rem;
}}

.month-chip {{
  font-family: var(--mono);
  font-size: 0.62rem;
  padding: 3px 8px;
  border-radius: 3px;
  border: 1px solid var(--border2);
  color: var(--text3);
}}

.month-chip.active {{ background: var(--gold2); border-color: var(--gold); color: var(--gold); }}
.month-chip.rising {{ background: var(--bg3); border-color: var(--border2); color: var(--text3); }}

/* ── EMPTY STATE ── */
.empty {{
  text-align: center;
  padding: 4rem 2rem;
  font-family: var(--mono);
  font-size: 0.75rem;
  color: var(--text3);
  letter-spacing: 0.05em;
}}

/* ── SCROLLBAR ── */
::-webkit-scrollbar {{ width: 6px; height: 6px; }}
::-webkit-scrollbar-track {{ background: var(--bg); }}
::-webkit-scrollbar-thumb {{ background: var(--border2); border-radius: 3px; }}
::-webkit-scrollbar-thumb:hover {{ background: var(--text3); }}

/* ── RESPONSIVE ── */
@media (max-width: 700px) {{
  header {{ padding: 0 1rem; }}
  .header-sub {{ display: none; }}
  nav {{ padding: 0 1rem; }}
  .catalogue-controls {{ padding: 0.8rem 1rem; top: 56px; }}
  .tab {{ padding: 0.6rem 0.8rem; font-size: 0.65rem; }}
  .col-type, .col-const, .col-rig, .col-months, .col-hours {{ display: none; }}
  .expand-content {{ grid-template-columns: 1fr; }}
  .expand-actions {{ align-items: flex-start; flex-direction: row; flex-wrap: wrap; }}
  .detail-meta-grid {{ grid-template-columns: 1fr 1fr; }}
  .planner-wrap {{ padding: 1rem; }}
}}
</style>
</head>
<body>

<header>
  <div class="header-left">
    <span class="header-title">The Southern 100</span>
    <span class="header-sub">Ocean Grove · Lat −38.16° · Bortle 4.5</span>
  </div>
  <div id="header-stats" style="font-family:var(--mono);font-size:0.62rem;color:var(--text3);text-align:right;line-height:1.8;"></div>
</header>

<nav>
  <button class="tab active" data-view="catalogue">Catalogue</button>
  <button class="tab" data-view="planner">Seasonal Planner</button>
</nav>

<!-- CATALOGUE VIEW -->
<div id="view-catalogue" class="view active">
  <div class="catalogue-controls">
    <div class="search-wrap">
      <svg width="14" height="14" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><circle cx="11" cy="11" r="8"/><path d="m21 21-4.35-4.35"/></svg>
      <input type="search" id="search" placeholder="Search name or catalogue ID…" autocomplete="off">
    </div>
    <select id="filter-status">
      <option value="">All status</option>
      <option>Complete</option>
      <option>Improve</option>
      <option>Partial</option>
      <option>To Image</option>
    </select>
    <select id="filter-type">
      <option value="">All types</option>
      <option>Dark Interstellar Medium</option>
      <option>Reflection Environments</option>
      <option>Cometary Globules</option>
      <option>Ionised Star-Forming Regions</option>
      <option>Open Clusters in Context</option>
      <option>Globular Clusters</option>
      <option>Stellar Feedback &amp; Shock Physics</option>
      <option>Planetary Nebulae</option>
      <option>External Galaxies</option>
      <option>Galaxy Groups &amp; Clusters</option>
      <option>Boundary / Mixed-Regime Objects</option>
    </select>
    <select id="filter-act">
      <option value="">All acts</option>
      <option value="I">Act I — Dark ISM</option>
      <option value="II">Act II — Reflection</option>
      <option value="III">Act III — Cometary Globules</option>
      <option value="IV">Act IV — HII Regions</option>
      <option value="V-A">Act V-A — Open Clusters</option>
      <option value="V-B">Act V-B — Globulars</option>
      <option value="VI">Act VI — Shock Physics</option>
      <option value="VII">Act VII — Planetary Nebulae</option>
      <option value="VIII">Act VIII — Galaxies</option>
      <option value="IX">Act IX — Galaxy Groups</option>
      <option value="Appendix">Appendix</option>
    </select>
    <select id="filter-month">
      <option value="">All months</option>
      <option>Jan</option><option>Feb</option><option>Mar</option>
      <option>Apr</option><option>May</option><option>Jun</option>
      <option>Jul</option><option>Aug</option><option>Sep</option>
      <option>Oct</option><option>Nov</option><option>Dec</option>
    </select>
    <select id="filter-rig">
      <option value="">All rigs</option>
      <option value="FRA300">FRA300</option>
      <option value="585MC">ASI585MC Pro</option>
      <option value="ATR2600c">ATR2600c</option>
    </select>
    <span class="filter-count" id="filter-count"></span>
  </div>

  <div class="table-wrap">
    <table id="cat-table">
      <thead>
        <tr>
          <th class="col-num" data-col="num">#<span class="sort-arrow">↕</span></th>
          <th data-col="name">Target<span class="sort-arrow">↕</span></th>
          <th class="col-type" data-col="type">Type<span class="sort-arrow">↕</span></th>
          <th class="col-const" data-col="constellation">Const.<span class="sort-arrow">↕</span></th>
          <th class="col-rig" data-col="best_rig">Rig<span class="sort-arrow">↕</span></th>
          <th class="col-months" data-col="best_months">Months<span class="sort-arrow">↕</span></th>
          <th class="col-hours" data-col="hours">Hours<span class="sort-arrow">↕</span></th>
          <th data-col="status">Status<span class="sort-arrow">↕</span></th>
        </tr>
      </thead>
      <tbody id="cat-body"></tbody>
    </table>
    <div class="empty" id="no-results" style="display:none">No targets match the current filters.</div>
  </div>
</div>

<!-- PLANNER VIEW -->
<div id="view-planner" class="view">
  <div class="planner-wrap">
    <div class="planner-title">Seasonal Visibility — Ocean Grove, Lat −38.16°&nbsp;&nbsp;·&nbsp;&nbsp;● Prime &nbsp; ↑ Rising/Setting</div>
    <table class="planner-table" id="planner-table">
      <thead>
        <tr>
          <th class="name-col">Target</th>
          <th>Jan</th><th>Feb</th><th>Mar</th><th>Apr</th><th>May</th><th>Jun</th>
          <th>Jul</th><th>Aug</th><th>Sep</th><th>Oct</th><th>Nov</th><th>Dec</th>
        </tr>
      </thead>
      <tbody id="planner-body"></tbody>
    </table>
  </div>
</div>

<!-- DETAIL VIEW -->
<div id="view-detail" class="view">
  <div class="detail-wrap">
    <button class="detail-back" id="detail-back">
      <svg width="12" height="12" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><path d="M19 12H5M12 5l-7 7 7 7"/></svg>
      Back to catalogue
    </button>
    <div id="detail-content"></div>
  </div>
</div>

<script>
const CATALOGUE = {data_json};

const MONTHS = ['Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec'];

// ── STATE ──
let sortCol = 'num';
let sortDir = 1;
let expandedNum = null;
let currentView = 'catalogue';

// ── UTILS ──
function statusClass(s) {{
  return 'status-' + s.replace(/\\s+/g, '-');
}}

function escHtml(s) {{
  return String(s)
    .replace(/&/g,'&amp;').replace(/</g,'&lt;').replace(/>/g,'&gt;')
    .replace(/"/g,'&quot;');
}}

function formatNotes(text) {{
  // Split on | and render as paragraphs
  return text.split('|').map(p => p.trim()).filter(Boolean)
    .map(p => `<p>${{escHtml(p)}}</p>`).join('');
}}

function formatCapSeq(text) {{
  // Split on | to get segments, detect labels (ALL CAPS followed by :)
  const segments = text.split('|').map(s => s.trim()).filter(Boolean);
  return segments.map(seg => {{
    const labelMatch = seg.match(/^([A-Z][A-Z0-9\\s\\/\\-\\+\\.″′°()]+):\\s*/);
    if (labelMatch) {{
      const label = labelMatch[1];
      const body = seg.slice(labelMatch[0].length);
      return `<div class="capseq-segment">
        <div class="capseq-label">${{escHtml(label)}}</div>
        <div class="capseq-text">${{escHtml(body)}}</div>
      </div>`;
    }}
    return `<div class="capseq-segment"><div class="capseq-text">${{escHtml(seg)}}</div></div>`;
  }}).join('');
}}

// ── HEADER STATS ──
function renderHeaderStats() {{
  const total = CATALOGUE.length;
  const complete = CATALOGUE.filter(e => e.status === 'Complete').length;
  const improve = CATALOGUE.filter(e => e.status === 'Improve').length;
  const partial = CATALOGUE.filter(e => e.status === 'Partial').length;
  document.getElementById('header-stats').innerHTML =
    `${{total}} targets &nbsp;·&nbsp; ${{complete}} complete &nbsp;·&nbsp; ${{improve}} improve &nbsp;·&nbsp; ${{partial}} partial`;
}}

// ── FILTER & SORT ──
function getFiltered() {{
  const q = document.getElementById('search').value.toLowerCase();
  const fStatus = document.getElementById('filter-status').value;
  const fType   = document.getElementById('filter-type').value;
  const fAct    = document.getElementById('filter-act').value;
  const fMonth  = document.getElementById('filter-month').value;
  const fRig    = document.getElementById('filter-rig').value;

  let data = CATALOGUE.filter(e => {{
    if (q && !e.name.toLowerCase().includes(q) && !e.cat_id.toLowerCase().includes(q)) return false;
    if (fStatus && e.status !== fStatus) return false;
    if (fType   && e.type !== fType) return false;
    if (fAct    && e.act !== fAct) return false;
    if (fMonth  && !(e.planner[fMonth] === '●' || e.planner[fMonth] === '↑')) return false;
    if (fRig    && !e.best_rig.includes(fRig)) return false;
    return true;
  }});

  data.sort((a, b) => {{
    let av = a[sortCol], bv = b[sortCol];
    if (sortCol === 'num') {{ av = +av; bv = +bv; }}
    else {{ av = String(av).toLowerCase(); bv = String(bv).toLowerCase(); }}
    return av < bv ? -sortDir : av > bv ? sortDir : 0;
  }});

  return data;
}}

// ── RENDER TABLE ──
function renderTable() {{
  const data = getFiltered();
  const tbody = document.getElementById('cat-body');
  const noRes = document.getElementById('no-results');
  document.getElementById('filter-count').textContent = `${{data.length}} / ${{CATALOGUE.length}}`;

  if (!data.length) {{
    tbody.innerHTML = '';
    noRes.style.display = '';
    return;
  }}
  noRes.style.display = 'none';

  let html = '';
  data.forEach(e => {{
    const isExp = expandedNum === e.num;
    html += `<tr data-num="${{e.num}}" class="${{isExp ? 'expanded' : ''}}">
      <td class="col-num">${{e.num}}</td>
      <td class="col-name">
        <strong>${{escHtml(e.name)}}</strong>
        <span>${{escHtml(e.cat_id)}}</span>
      </td>
      <td class="col-type">${{escHtml(e.type)}}</td>
      <td class="col-const">${{escHtml(e.constellation)}}</td>
      <td class="col-rig">${{escHtml(e.best_rig)}}</td>
      <td class="col-months">${{escHtml(e.best_months)}}</td>
      <td class="col-hours">${{escHtml(e.hours)}}</td>
      <td><span class="status-badge ${{statusClass(e.status)}}">${{escHtml(e.status)}}</span></td>
    </tr>`;
    if (isExp) {{
      html += `<tr class="expand-row"><td colspan="8">
        <div class="expand-content">
          <div>
            <div class="expand-meta">
              <span class="meta-item"><span class="meta-label">RA</span><span class="meta-val">${{escHtml(e.ra)}}</span></span>
              <span class="meta-item"><span class="meta-label">Dec</span><span class="meta-val">${{escHtml(e.dec)}}</span></span>
              <span class="meta-item"><span class="meta-label">Rotation</span><span class="meta-val">${{escHtml(e.rotation)}}</span></span>
              <span class="meta-item"><span class="meta-label">Size</span><span class="meta-val">${{escHtml(e.size)}}′</span></span>
              <span class="meta-item"><span class="meta-label">Transit</span><span class="meta-val">${{escHtml(e.transit_alt)}}</span></span>
              <span class="meta-item"><span class="meta-label">Circumpolar</span><span class="meta-val">${{escHtml(e.circumpolar)}}</span></span>
              <span class="meta-item"><span class="meta-label">Filter</span><span class="meta-val">${{escHtml(e.filter_rec)}}</span></span>
              <span class="meta-item"><span class="meta-label">Moon sep</span><span class="meta-val">${{escHtml(e.moon_sep)}}°</span></span>
            </div>
            <div class="notes-text">${{formatNotes(e.notes)}}</div>
          </div>
          <div class="expand-actions">
            <button class="btn btn-primary" onclick="showDetail(${{e.num}});event.stopPropagation()">Capture sequence →</button>
          </div>
        </div>
      </td></tr>`;
    }}
  }});

  tbody.innerHTML = html;

  // row click
  tbody.querySelectorAll('tr[data-num]').forEach(row => {{
    row.addEventListener('click', () => {{
      const num = +row.dataset.num;
      expandedNum = expandedNum === num ? null : num;
      renderTable();
    }});
  }});
}}

// ── RENDER PLANNER ──
function renderPlanner() {{
  const tbody = document.getElementById('planner-body');
  let html = '';
  CATALOGUE.forEach(e => {{
    html += `<tr data-num="${{e.num}}"><td class="name-col"><span class="p-num">#${{e.num}}</span><span class="p-name">${{escHtml(e.name)}}</span></td>`;
    MONTHS.forEach(m => {{
      const v = e.planner[m] || '';
      if (v === '●') html += `<td><span class="p-dot">●</span></td>`;
      else if (v === '↑') html += `<td><span class="p-rise">↑</span></td>`;
      else html += `<td></td>`;
    }});
    html += '</tr>';
  }});
  tbody.innerHTML = html;

  tbody.querySelectorAll('tr[data-num]').forEach(row => {{
    row.addEventListener('click', () => {{
      showDetail(+row.dataset.num);
    }});
  }});
}}

// ── DETAIL VIEW ──
function showDetail(num) {{
  const e = CATALOGUE.find(x => x.num === num);
  if (!e) return;

  const monthChips = MONTHS.map(m => {{
    const v = e.planner[m] || '';
    const cls = v === '●' ? 'active' : v === '↑' ? 'rising' : '';
    return `<span class="month-chip ${{cls}}">${{m}}</span>`;
  }}).join('');

  document.getElementById('detail-content').innerHTML = `
    <div class="detail-header">
      <div class="detail-num">Act ${{escHtml(e.act)}} · #${{e.num}}</div>
      <div class="detail-name">${{escHtml(e.name)}}</div>
      <div class="detail-catid">${{escHtml(e.cat_id)}}</div>
    </div>

    <div class="detail-meta-grid">
      <div class="detail-meta-item"><span class="detail-meta-label">Type</span><span class="detail-meta-val">${{escHtml(e.type)}}</span></div>
      <div class="detail-meta-item"><span class="detail-meta-label">Constellation</span><span class="detail-meta-val">${{escHtml(e.constellation)}}</span></div>
      <div class="detail-meta-item"><span class="detail-meta-label">RA</span><span class="detail-meta-val">${{escHtml(e.ra)}}</span></div>
      <div class="detail-meta-item"><span class="detail-meta-label">Dec</span><span class="detail-meta-val">${{escHtml(e.dec)}}</span></div>
      <div class="detail-meta-item"><span class="detail-meta-label">Rotation</span><span class="detail-meta-val">${{escHtml(e.rotation)}}</span></div>
      <div class="detail-meta-item"><span class="detail-meta-label">Size</span><span class="detail-meta-val">${{escHtml(e.size)}}′</span></div>
      <div class="detail-meta-item"><span class="detail-meta-label">Transit Alt</span><span class="detail-meta-val">${{escHtml(e.transit_alt)}}</span></div>
      <div class="detail-meta-item"><span class="detail-meta-label">Circumpolar</span><span class="detail-meta-val">${{escHtml(e.circumpolar)}}</span></div>
      <div class="detail-meta-item"><span class="detail-meta-label">Best Rig</span><span class="detail-meta-val">${{escHtml(e.best_rig)}}</span></div>
      <div class="detail-meta-item"><span class="detail-meta-label">Filter</span><span class="detail-meta-val">${{escHtml(e.filter_rec)}}</span></div>
      <div class="detail-meta-item"><span class="detail-meta-label">Hours</span><span class="detail-meta-val">${{escHtml(e.hours)}}</span></div>
      <div class="detail-meta-item"><span class="detail-meta-label">Moon Sep</span><span class="detail-meta-val">${{escHtml(e.moon_sep)}}°</span></div>
      <div class="detail-meta-item"><span class="detail-meta-label">Status</span><span class="detail-meta-val"><span class="status-badge ${{statusClass(e.status)}}">${{escHtml(e.status)}}</span></span></div>
    </div>

    <div class="detail-section-title">Seasonal Visibility</div>
    <div class="detail-planner">${{monthChips}}</div>

    <div class="detail-section-title">Notes</div>
    <div class="detail-notes">${{formatNotes(e.notes)}}</div>

    <div class="detail-section-title">Capture Sequence</div>
    <div class="detail-capseq">${{formatCapSeq(e.capture_seq)}}</div>
  `;

  setView('detail');
}}

// ── VIEW SWITCHING ──
function setView(view) {{
  currentView = view;
  document.querySelectorAll('.view').forEach(v => v.classList.remove('active'));
  document.querySelectorAll('.tab').forEach(t => t.classList.remove('active'));

  if (view === 'detail') {{
    document.getElementById('view-detail').classList.add('active');
  }} else {{
    document.getElementById(`view-${{view}}`).classList.add('active');
    document.querySelector(`.tab[data-view="${{view}}"]`).classList.add('active');
  }}
}}

// ── SORT ──
function setSort(col) {{
  if (sortCol === col) sortDir *= -1;
  else {{ sortCol = col; sortDir = 1; }}
  document.querySelectorAll('thead th').forEach(th => {{
    th.classList.toggle('sorted', th.dataset.col === col);
    if (th.dataset.col === col) {{
      th.querySelector('.sort-arrow').textContent = sortDir === 1 ? '↓' : '↑';
    }} else {{
      th.querySelector('.sort-arrow').textContent = '↕';
    }}
  }});
  renderTable();
}}

// ── INIT ──
document.querySelectorAll('.tab').forEach(tab => {{
  tab.addEventListener('click', () => setView(tab.dataset.view));
}});

document.getElementById('detail-back').addEventListener('click', () => setView('catalogue'));

['search','filter-status','filter-type','filter-act','filter-month','filter-rig'].forEach(id => {{
  document.getElementById(id).addEventListener('input', () => {{ expandedNum = null; renderTable(); }});
}});

document.querySelectorAll('thead th[data-col]').forEach(th => {{
  th.addEventListener('click', () => setSort(th.dataset.col));
}});

renderHeaderStats();
renderTable();
renderPlanner();
</script>
</body>
</html>'''
    return html

if __name__ == '__main__':
    print(f"Reading {XLSX_PATH}...")
    entries = extract_data(XLSX_PATH)
    print(f"  {len(entries)} entries extracted")
    print(f"Building HTML...")
    html = build_html(entries)
    with open(OUTPUT_PATH, 'w', encoding='utf-8') as f:
        f.write(html)
    size_kb = os.path.getsize(OUTPUT_PATH) / 1024
    print(f"  Written to {OUTPUT_PATH} ({size_kb:.0f} KB)")
    print("Done.")
